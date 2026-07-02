import argparse
import json
import shutil
import subprocess
import sys
import time
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def resolve_cmd(cmd):
    if not cmd:
        return cmd
    resolved = shutil.which(cmd[0])
    if resolved:
        return [resolved, *cmd[1:]]
    return cmd


def run(cmd, cwd=None, timeout=300):
    proc = subprocess.run(
        resolve_cmd(cmd),
        cwd=str(cwd) if cwd else None,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        timeout=timeout,
        shell=False,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            "Command failed:\n"
            + " ".join(map(str, cmd))
            + "\nSTDOUT:\n"
            + proc.stdout
            + "\nSTDERR:\n"
            + proc.stderr
        )
    return proc.stdout.strip()


def first_json_object(text):
    start = text.find("{")
    if start < 0:
        raise ValueError("No JSON object found in command output.")
    depth = 0
    in_str = False
    escape = False
    for i, ch in enumerate(text[start:], start):
        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return json.loads(text[start : i + 1])
    raise ValueError("JSON object was not balanced.")


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=True, indent=2), encoding="utf-8")


def load_config(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def output_paths(cfg):
    root = Path(cfg["output_root"])
    script = cfg["script_id"]
    media_root = root / "04_逐镜头视频图片下载" / script
    table_root = root / "05_带视频图片表格" / script
    return {
        "media_root": media_root,
        "image_dir": media_root / "images",
        "video_dir": media_root / "videos",
        "manifest": media_root / f"{script}_视频图片_manifest.json",
        "table_root": table_root,
        "upload_xlsx": table_root / f"{script}_拍摄脚本分镜_1_飞书上传版.xlsx",
        "uploads": table_root / f"{script}_feishu_video_drive_uploads.json",
        "video_cells": table_root / f"{script}_feishu_video_richtext_cells.json",
        "permission_data": table_root / "permission_data_nobom.json",
    }


def extract_first_url(node_data):
    url = node_data.get("data", {}).get("url")
    if isinstance(url, list) and url:
        return url[0]
    if isinstance(url, str) and url:
        return url
    return ""


def collect_manifest(cfg):
    manifest = []
    for node in sorted(cfg["nodes"], key=lambda x: int(x["shot"])):
        img = first_json_object(
            run(["libtv", "node", "-p", cfg["project_uuid"], node["image_node"]], timeout=60)
        )
        vid = first_json_object(
            run(["libtv", "node", "-p", cfg["project_uuid"], node["video_node"]], timeout=60)
        )
        image_url = extract_first_url(img)
        video_url = extract_first_url(vid)
        if not image_url or not video_url:
            raise RuntimeError(f"Shot {node['shot']} missing image/video url.")
        manifest.append(
            {
                "shot": int(node["shot"]),
                "imageNode": node["image_node"],
                "videoNode": node["video_node"],
                "imageUrl": image_url,
                "videoUrl": video_url,
                "durationSetting": vid["data"]["params"]["settings"]["duration"],
                "taskId": vid["data"].get("taskInfo", {}).get("taskId", ""),
            }
        )
    paths = output_paths(cfg)
    write_json(paths["manifest"], manifest)
    print(f"[collect] wrote {paths['manifest']}")
    return manifest


def load_manifest(cfg):
    paths = output_paths(cfg)
    if not paths["manifest"].exists():
        return collect_manifest(cfg)
    return json.loads(paths["manifest"].read_text(encoding="utf-8-sig"))


def download_media(cfg):
    paths = output_paths(cfg)
    paths["image_dir"].mkdir(parents=True, exist_ok=True)
    paths["video_dir"].mkdir(parents=True, exist_ok=True)
    manifest = load_manifest(cfg)
    for item in manifest:
        shot = f"{int(item['shot']):02d}"
        image_path = paths["image_dir"] / f"{cfg['script_id']}_镜头{shot}_分镜图.png"
        video_path = paths["video_dir"] / f"{cfg['script_id']}_镜头{shot}_视频.mp4"
        if not image_path.exists():
            urllib.request.urlretrieve(item["imageUrl"], image_path)
        if not video_path.exists():
            urllib.request.urlretrieve(item["videoUrl"], video_path)
        item["imagePath"] = str(image_path)
        item["videoPath"] = str(video_path)
    write_json(paths["manifest"], manifest)
    print(f"[download] images={len(list(paths['image_dir'].glob('*.png')))} videos={len(list(paths['video_dir'].glob('*.mp4')))}")
    return manifest


def build_upload_xlsx(cfg):
    paths = output_paths(cfg)
    paths["table_root"].mkdir(parents=True, exist_ok=True)
    manifest = load_manifest(cfg)
    by_shot = {int(x["shot"]): x for x in manifest}
    source = Path(cfg["source_xlsx"])
    shutil.copy2(source, paths["upload_xlsx"])

    wb = load_workbook(paths["upload_xlsx"])
    ws = wb[wb.sheetnames[0]]
    headers = ["生成时长(秒)", "分镜图", "分镜图链接", "视频链接", "视频任务ID"]
    start_col = ws.max_column + 1
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="D9E2F3")

    for offset, header in enumerate(headers):
        cell = ws.cell(row=1, column=start_col + offset, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="1F1F1F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in range(2, ws.max_row + 1):
        shot_text = str(ws.cell(row=row, column=3).value or "")
        digits = "".join(ch for ch in shot_text if ch.isdigit())
        if not digits:
            continue
        item = by_shot.get(int(digits))
        if not item:
            continue
        ws.cell(row=row, column=start_col, value=item["durationSetting"])
        image_path = Path(item["imagePath"])
        if image_path.exists():
            img = XLImage(str(image_path))
            img.width = 72
            img.height = 128
            ws.add_image(img, ws.cell(row=row, column=start_col + 1).coordinate)
        img_cell = ws.cell(row=row, column=start_col + 2, value="查看分镜图")
        img_cell.hyperlink = item["imageUrl"]
        img_cell.style = "Hyperlink"
        vid_cell = ws.cell(row=row, column=start_col + 3, value="查看视频")
        vid_cell.hyperlink = item["videoUrl"]
        vid_cell.style = "Hyperlink"
        ws.cell(row=row, column=start_col + 4, value=item.get("taskId", ""))
        ws.row_dimensions[row].height = 100
        for col in range(start_col, start_col + len(headers)):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    widths = [12, 14, 14, 12, 24]
    for offset, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + offset)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(paths["upload_xlsx"])
    print(f"[xlsx] wrote {paths['upload_xlsx']}")
    return paths["upload_xlsx"]


def import_sheet(cfg):
    paths = output_paths(cfg)
    if not paths["upload_xlsx"].exists():
        build_upload_xlsx(cfg)
    out = run(
        [
            "lark-cli",
            "drive",
            "+import",
            "--file",
            f".\\{paths['upload_xlsx'].name}",
            "--type",
            "sheet",
            "--name",
            cfg["feishu_sheet_name"],
            "--as",
            "user",
            "--format",
            "json",
        ],
        cwd=paths["upload_xlsx"].parent,
        timeout=600,
    )
    data = first_json_object(out)
    token = data["data"]["token"]
    print(f"[import] {data['data']['url']}")
    return token


def upload_videos(cfg):
    paths = output_paths(cfg)
    manifest = load_manifest(cfg)
    uploads = []
    for item in sorted(manifest, key=lambda x: int(x["shot"])):
        video_path = Path(item["videoPath"])
        out = run(
            [
                "lark-cli",
                "drive",
                "+upload",
                "--file",
                f".\\{video_path.name}",
                "--name",
                video_path.name,
                "--as",
                "user",
                "--format",
                "json",
            ],
            cwd=video_path.parent,
            timeout=600,
        )
        data = first_json_object(out)
        uploads.append(
            {
                "shot": int(item["shot"]),
                "name": video_path.name,
                "token": data["data"]["file_token"],
                "url": data["data"]["url"],
            }
        )
        print(f"[upload-video] {video_path.name} -> {data['data']['url']}")
        time.sleep(1)
    write_json(paths["uploads"], uploads)
    return uploads


def load_uploads(cfg):
    paths = output_paths(cfg)
    if not paths["uploads"].exists():
        return upload_videos(cfg)
    return json.loads(paths["uploads"].read_text(encoding="utf-8-sig"))


def patch_video_links(cfg, spreadsheet_token):
    paths = output_paths(cfg)
    uploads = sorted(load_uploads(cfg), key=lambda x: int(x["shot"]))
    cells = [
        [
            {
                "rich_text": [{"type": "link", "text": "查看视频", "link": item["url"]}],
                "cell_styles": {
                    "font_color": "#0000FF",
                    "horizontal_alignment": "center",
                    "vertical_alignment": "middle",
                },
            }
        ]
        for item in uploads
    ]
    write_json(paths["video_cells"], cells)
    col = cfg.get("video_link_column", "N")
    end_row = len(cells) + 1
    run(
        [
            "lark-cli",
            "sheets",
            "+cells-set",
            "--spreadsheet-token",
            spreadsheet_token,
            "--sheet-name",
            cfg.get("sheet_name", "分镜脚本"),
            "--range",
            f"{col}2:{col}{end_row}",
            "--cells",
            f"@{paths['video_cells'].name}",
            "--as",
            "user",
            "--format",
            "json",
        ],
        cwd=paths["video_cells"].parent,
        timeout=120,
    )
    print(f"[patch-links] updated {col}2:{col}{end_row}")


def set_permissions(cfg, spreadsheet_token):
    paths = output_paths(cfg)
    uploads = load_uploads(cfg)
    data = {
        "link_share_entity": "tenant_readable",
        "share_entity": "same_tenant",
        "security_entity": "anyone_can_view",
    }
    write_json(paths["permission_data"], data)

    def patch(token, doc_type, name):
        params_file = paths["table_root"] / f"permission_{doc_type}_{token}.json"
        write_json(params_file, {"token": token, "type": doc_type})
        run(
            [
                "lark-cli",
                "drive",
                "permission.public",
                "patch",
                "--params",
                f"@{params_file.name}",
                "--data",
                f"@{paths['permission_data'].name}",
                "--as",
                "user",
                "--yes",
                "--format",
                "json",
            ],
            cwd=paths["table_root"],
            timeout=120,
        )
        print(f"[permission] {name}")
        time.sleep(0.8)

    patch(spreadsheet_token, "sheet", "sheet")
    for item in uploads:
        patch(item["token"], "file", item["name"])


def verify(cfg, spreadsheet_token):
    info = first_json_object(
        run(
            [
                "lark-cli",
                "sheets",
                "+workbook-info",
                "--spreadsheet-token",
                spreadsheet_token,
                "--as",
                "user",
                "--format",
                "json",
            ],
            timeout=60,
        )
    )
    print("[verify] workbook sheets:", info["data"]["sheets"])
    col = cfg.get("video_link_column", "N")
    cells = first_json_object(
        run(
            [
                "lark-cli",
                "sheets",
                "+cells-get",
                "--spreadsheet-token",
                spreadsheet_token,
                "--sheet-name",
                cfg.get("sheet_name", "分镜脚本"),
                "--range",
                f"{col}2:{col}3",
                "--as",
                "user",
                "--format",
                "json",
            ],
            timeout=60,
        )
    )
    print("[verify] sample video cells:", cells["data"]["ranges"][0]["cells"])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True)
    parser.add_argument(
        "--steps",
        default="collect,download,xlsx,import,upload-videos,patch-links,permissions,verify",
        help="Comma-separated steps.",
    )
    parser.add_argument("--spreadsheet-token", default="")
    args = parser.parse_args()

    cfg = load_config(args.config)
    steps = [s.strip() for s in args.steps.split(",") if s.strip()]
    token = args.spreadsheet_token

    for step in steps:
        if step == "collect":
            collect_manifest(cfg)
        elif step == "download":
            download_media(cfg)
        elif step == "xlsx":
            build_upload_xlsx(cfg)
        elif step == "import":
            token = import_sheet(cfg)
        elif step == "upload-videos":
            upload_videos(cfg)
        elif step == "patch-links":
            if not token:
                raise SystemExit("--spreadsheet-token is required unless import ran in the same command.")
            patch_video_links(cfg, token)
        elif step == "permissions":
            if not token:
                raise SystemExit("--spreadsheet-token is required unless import ran in the same command.")
            set_permissions(cfg, token)
        elif step == "verify":
            if not token:
                raise SystemExit("--spreadsheet-token is required unless import ran in the same command.")
            verify(cfg, token)
        else:
            raise SystemExit(f"Unknown step: {step}")

    if token:
        print(f"[done] https://ucnscwivbsrz.feishu.cn/sheets/{token}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"[error] {exc}", file=sys.stderr)
        raise
