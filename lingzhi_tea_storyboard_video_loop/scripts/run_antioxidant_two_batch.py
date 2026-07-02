import json
import re
import subprocess
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl


LIBTV = r"C:\Users\gba\.libtv\libtv.exe"
PROJECT_UUID = "4bc1c28bb3754d8ca3521aa6df975130"
STATE_PATH = Path(r"E:\灵鹤芝谷素材库\灵鹤芝谷工具矩阵\outputs\批量脚本视频任务_0623\antioxidant_two_run_state.json")
DYNAMIC_TASKS_PATH = Path(r"E:\灵鹤芝谷素材库\灵鹤芝谷工具矩阵\outputs\批量脚本视频任务_0623\legacy_dynamic_tasks.json")

JUNK_MARKERS = ["????", "\ufffd", "锟斤拷"]
SKIP_VIDEO = set()
STATE_LOCK = threading.Lock()

TASKS = {
    "antioxidant1": {
        "source_xlsx": "C:\\Users\\gba\\Downloads\\\u65b0\u7075\u829d\u8336\u62cd\u6444\u811a\u672c\u5206\u955c-\u6297\u6c27\u5316-1.xlsx",
        "group_id": "fa89695d-9bff-488c-91f8-c4d3514ac649",
        "baseline": "efc70b18-35a9-4ff4-ba59-95dc4c4dfbbe",
        "product_refs": [
            "e8b76923-dd86-4e36-862a-8da2b3aea089",
            "2e3fdaaa-050d-4b0b-950e-e144b216b378",
            "d706a387-31cf-4210-8d84-8b983441918d",
        ],
    },
    "antioxidant4": {
        "source_xlsx": "C:\\Users\\gba\\Downloads\\\u65b0\u7075\u829d\u8336\u62cd\u6444\u811a\u672c\u5206\u955c-\u6297\u6c27\u5316-4.xlsx",
        "group_id": "17493ee7-697a-4a73-9eb4-059b7243c91e",
        "baseline": "671d10ae-e2c3-4d56-90ee-97498e259618",
        "product_refs": [
            "b8c2b970-20c0-4f70-b817-63c3ae0b06f5",
            "3503a613-d6be-4d6f-a353-1bdaae04eff9",
            "ae7eca29-6f4e-4163-9a9e-05ffe598049d",
        ],
    },
    "script02": {
        "source_xlsx": "E:/\u7075\u9e64\u829d\u8c37\u7d20\u6750\u5e93/\u7075\u9e64\u829d\u8c37\u5de5\u5177\u77e9\u9635/outputs/\u7075\u829d\u833624\u7ec4_\u5206\u955c\u56fe\u7247\u89c6\u9891\u6d41\u7a0b/01_\u62c6\u5206\u811a\u672c/\u811a\u672c02/\u811a\u672c02_\u62cd\u6444\u811a\u672c\u5206\u955c_2.xlsx",
        "group_id": "24d32565-dec0-4b39-8ccf-4f1e225874e4",
        "baseline": "efc70b18-35a9-4ff4-ba59-95dc4c4dfbbe",
        "product_refs": [
            "e8b76923-dd86-4e36-862a-8da2b3aea089",
            "2e3fdaaa-050d-4b0b-950e-e144b216b378",
            "d706a387-31cf-4210-8d84-8b983441918d",
        ],
    },
}

if DYNAMIC_TASKS_PATH.exists():
    TASKS.update(json.loads(DYNAMIC_TASKS_PATH.read_text(encoding="utf-8")))


def run(cmd, timeout=300):
    proc = subprocess.run(
        cmd,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        timeout=timeout,
        shell=False,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            "Command failed:\n"
            + " ".join(cmd)
            + "\nSTDOUT:\n"
            + proc.stdout
            + "\nSTDERR:\n"
            + proc.stderr
        )
    return proc.stdout


def run_allow_created(cmd, timeout=300):
    proc = subprocess.run(
        cmd,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        timeout=timeout,
        shell=False,
    )
    if proc.returncode != 0 and proc.stdout.find("{") < 0:
        raise RuntimeError(
            "Command failed:\n"
            + " ".join(cmd)
            + "\nSTDOUT:\n"
            + proc.stdout
            + "\nSTDERR:\n"
            + proc.stderr
        )
    return proc.stdout


def first_json(text):
    start = text.find("{")
    if start < 0:
        raise ValueError(text[:1000])
    depth = 0
    in_str = False
    esc = False
    for i, ch in enumerate(text[start:], start):
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return json.loads(text[start : i + 1])
    raise ValueError("unbalanced json")


def load_state():
    with STATE_LOCK:
        if STATE_PATH.exists():
            text = STATE_PATH.read_text(encoding="utf-8")
            if text.strip():
                return json.loads(text)
        return {"tasks": {}}


def save_state(state):
    with STATE_LOCK:
        STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
        tmp = STATE_PATH.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(STATE_PATH)


def update_shot_state(task_name, shot_key, values):
    with STATE_LOCK:
        if STATE_PATH.exists():
            text = STATE_PATH.read_text(encoding="utf-8")
            state = json.loads(text) if text.strip() else {"tasks": {}}
        else:
            state = {"tasks": {}}
        item = state["tasks"].setdefault(task_name, {}).setdefault(str(shot_key), {})
        item.update(values)
        STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
        tmp = STATE_PATH.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(STATE_PATH)


def read_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    if str(ws.cell(1, 1).value or "") == "脚本编号":
        for r in range(2, ws.max_row + 1):
            shot_raw = str(ws.cell(r, 3).value or "")
            digits = "".join(ch for ch in shot_raw if ch.isdigit())
            visual = ws.cell(r, 7).value
            line = ws.cell(r, 8).value
            if not digits and not visual and not line:
                continue
            rows.append(
                {
                    "row": r,
                    "shot": int(digits),
                    "time": str(ws.cell(r, 4).value or ""),
                    "visual": str(visual or ""),
                    "action": str(ws.cell(r, 7).value or ""),
                    "scene": str(ws.cell(r, 7).value or ""),
                    "shot_size": str(ws.cell(r, 5).value or ""),
                    "line": str(line or ""),
                    "bgm": str(ws.cell(r, 9).value or ""),
                }
            )
        return rows
    for r in range(3, ws.max_row + 1):
        shot = ws.cell(r, 1).value
        visual = ws.cell(r, 3).value
        line = ws.cell(r, 7).value
        if shot is None and not visual and not line:
            continue
        rows.append(
            {
                "row": r,
                "shot": int(float(shot)),
                "time": str(ws.cell(r, 2).value or ""),
                "visual": str(visual or ""),
                "action": str(ws.cell(r, 4).value or ""),
                "scene": str(ws.cell(r, 5).value or ""),
                "shot_size": str(ws.cell(r, 6).value or ""),
                "line": str(line or ""),
                "bgm": str(ws.cell(r, 8).value or ""),
            }
        )
    return rows


def estimate_duration(line, original_time):
    # Chinese voiceover normally lands around 4.5 chars/sec in generated UGC.
    chars = len(re.sub(r"\s|[，。！？、,.!?]", "", line or ""))
    if chars:
        sec = max(4, min(10, round(chars / 4.3 + 0.8)))
    else:
        nums = [int(x) for x in re.findall(r"\d+", original_time or "")]
        sec = max(nums) if nums else 4
    return int(sec)


def product_needed(row):
    text = row["visual"] + row["line"]
    keys = ["灵芝茶", "产品", "包装", "茶包", "这一包", "下单", "试喝", "盒"]
    return any(k in text for k in keys)


def subject_en(text):
    parts = []
    if "果蔬" in text or "番茄" in text or "紫甘蓝" in text or "彩椒" in text:
        parts.append("quick-cut closeups of fresh vegetables and fruits such as tomatoes, strawberries, kiwi, cucumber, purple cabbage and bell pepper")
    if "灵芝切片" in text or "灵芝片" in text:
        parts.append("macro closeup of sliced reishi mushroom texture on a clean table")
    if "九种食材" in text or "红枣" in text or "枸杞" in text or "黄芪" in text:
        parts.append("overhead closeup of nine tea ingredients in a white porcelain dish: jujube, goji berries, astragalus, codonopsis, longan, dried citrus peel, rose petals, jasmine and reishi pieces")
    if "拆开茶包" in text or "倒出原料" in text:
        parts.append("hands open a tea bag and pour the loose ingredients into a white porcelain dish, then gently touch the ingredients")
    if "产品包装" in text or "包装盒" in text or "产品盒" in text:
        parts.append("the Lingzhi Huangqi Gouqi Tea outer box sits naturally on the table, hand gently touches the box edge, packaging identity preserved from reference")
    if "茶包" in text and "拆开" not in text:
        parts.append("the creator holds or presents a white tea bag from the referenced tea product")
    if "茶杯" in text or "轻抿" in text or "闻" in text:
        parts.append("the creator holds a warm tea cup, smells or sips the tea with a relaxed smile")
    if "办公桌" in text:
        parts.append("the creator works at a desk with a laptop, the tea product placed nearby as a daily companion")
    if "镜子" in text or "脸颊" in text:
        parts.append("the creator looks in a mirror and gently touches her cheek, showing a fresh natural complexion")
    if "朋友" in text:
        parts.append("the creator chats with a friend in the living room; the friend looks pleasantly surprised while the creator smiles")
    if "挥手" in text or "赶紧冲" in text:
        parts.append("the creator waves to camera in a casual closing moment with the product visible nearby")
    if "女主" in text and not parts:
        parts.append("the same female creator speaks naturally to camera with relaxed posture and sincere eye contact")
    return "; ".join(parts) or "a natural daily UGC moment matching the script action"


def shot_size_en(text):
    if "特写" in text:
        return "close-up or macro detail shot"
    if "近景" in text:
        return "close shot"
    if "中景" in text:
        return "medium shot"
    if "俯拍" in text:
        return "overhead tabletop angle"
    return "natural handheld smartphone framing"


def motion_en(text):
    if "推" in text:
        return "slow natural push-in"
    if "扫" in text or "平移" in text:
        return "smooth lateral scan"
    if "快切" in text or "快速" in text:
        return "short-video quick-cut rhythm"
    if "定格" in text:
        return "clean final still-frame composition"
    return "slight handheld UGC camera movement"


def image_prompt(row, task_name):
    product_lock = (
        "Use the uploaded product references only for product consistency. "
        "The product must be Lingzhi Huangqi Gouqi Tea: kraft paper outer box, pink label window and white tea bag. "
        "Preserve the original printed Chinese packaging text and label placement exactly when visible. "
        "Do not turn it into Wuzhi Tea or any other product. "
        if product_needed(row)
        else "This shot does not need a front-facing product packshot; keep the same UGC creator and warm home style when a person appears. "
    )
    line_note = "The creator may be naturally speaking the shot voiceover with believable mouth shape, but do not design any overlay captions."
    return (
        f"9:16 vertical realistic UGC smartphone image, {shot_size_en(row['shot_size'] + row['visual'])}, "
        f"{motion_en(row['shot_size'] + row['visual'])}. "
        f"Script shot {row['shot']}: {subject_en(row['visual'] + row['action'])}. "
        f"Action and expression: {subject_en(row['action'])}; sincere relaxed female creator when present. "
        f"Scene: warm natural home or tabletop setting matching: {subject_en(row['scene'])}. "
        "Soft natural warm light, authentic phone-camera framing, realistic skin texture, not a polished commercial poster. "
        "Keep the same female creator identity, hair, outfit, age and environment from the baseline reference when a person appears. "
        + product_lock
        + line_note
    )


def video_prompt(row, duration):
    oral = row["line"]
    return (
        f"Create one independent {duration}-second vertical 9:16 UGC short-video shot from the left storyboard image. "
        f"Match this script shot exactly: {subject_en(row['visual'] + row['action'])}. "
        f"Camera movement: {motion_en(row['shot_size'] + row['visual'])}; shot size: {shot_size_en(row['shot_size'])}. "
        "Keep action, hand movement, product position, environment and creator identity consistent with the input image. "
        "Use natural handheld smartphone motion, realistic body mechanics, natural eye movement and relaxed facial expressions. "
        f"Voiceover line must be spoken completely and naturally within about {duration} seconds: {oral} "
        "Do not rewrite, shorten, expand or replace the voiceover line. "
        "No subtitles, no on-screen script text, no title text, no corner labels, no sticker copy and no watermark. "
        "Preserve original printed packaging text from the reference product when visible; no product appearance drift."
    )


def create_image(task_name, cfg, row):
    state = load_state()
    task_state = state["tasks"].setdefault(task_name, {})
    shot_key = str(row["shot"])
    shot_state = task_state.setdefault(shot_key, {})
    if shot_state.get("image_node"):
        return task_name, row["shot"], shot_state["image_node"], "exists"

    cmd = [
        LIBTV,
        "node",
        "create",
        f"{task_name}_shot{row['shot']:02d}_image",
        "-t",
        "image",
        "--prompt",
        image_prompt(row, task_name),
        "-s",
        "model=Lib Image",
        "-s",
        "modeType=image2image",
        "-s",
        "ratio=9:16",
        "-s",
        "quality=high",
        "-s",
        "resolution=2K",
        "-s",
        "count=1",
        "--left-add",
        cfg["baseline"],
    ]
    if product_needed(row):
        for ref in cfg["product_refs"]:
            cmd.extend(["--left-add", ref])
    out = run_allow_created(cmd, timeout=300)
    data = first_json(out)
    node_id = data.get("newNodeKey") or data.get("nodeKey") or data.get("data", {}).get("nodeKey") or data.get("id")
    if not node_id:
        # Some libtv versions return the node under data.key.
        node_id = data.get("data", {}).get("key")
    if not node_id:
        raise RuntimeError(f"Could not parse node id for {task_name} shot {row['shot']}: {out[:1000]}")
    run_allow_created([LIBTV, "group", cfg["group_id"], "--node", node_id], timeout=120)

    detail = first_json(run([LIBTV, "node", "-p", PROJECT_UUID, node_id], timeout=90))
    prompt = json.dumps(detail.get("data", {}).get("params", {}), ensure_ascii=False)
    if any(marker in prompt for marker in JUNK_MARKERS):
        raise RuntimeError(f"Prompt encoding failed for {task_name} shot {row['shot']} node {node_id}")

    update_shot_state(
        task_name,
        shot_key,
        {
            "image_node": node_id,
            "duration": estimate_duration(row["line"], row["time"]),
            "line": row["line"],
        },
    )
    return task_name, row["shot"], node_id, "created"


def scan_groups_to_state():
    state = load_state()
    rows_by_task = {name: read_rows(cfg["source_xlsx"]) for name, cfg in TASKS.items()}
    for task_name, cfg in TASKS.items():
        task_state = state["tasks"].setdefault(task_name, {})
        for row in rows_by_task[task_name]:
            shot_key = str(row["shot"])
            task_state.setdefault(shot_key, {})["duration"] = estimate_duration(row["line"], row["time"])
            task_state[shot_key]["line"] = row["line"]
        group = first_json(run([LIBTV, "group", cfg["group_id"]], timeout=90))
        ids = list(dict.fromkeys(group.get("executionPlan", []) + group.get("childNodeIds", [])))
        for node_id in ids:
            try:
                detail = first_json(run([LIBTV, "node", "-p", PROJECT_UUID, node_id], timeout=90))
            except Exception:
                continue
            data = detail.get("data", {})
            name = data.get("displayName") or data.get("name") or ""
            m = re.match(rf"{re.escape(task_name)}_shot(\d+)_image$", name)
            if not m:
                continue
            shot_key = str(int(m.group(1)))
            task_state.setdefault(shot_key, {})["image_node"] = node_id
            task_state[shot_key]["duration"] = task_state.get(shot_key, {}).get("duration") or 4
    save_state(state)
    return state


def image_status(node_id):
    data = first_json(run([LIBTV, "node", "-p", PROJECT_UUID, node_id], timeout=90))
    d = data.get("data", {})
    url = d.get("url")
    url_count = len(url) if isinstance(url, list) else (1 if url else 0)
    return {
        "loading": d.get("loading") or d.get("taskInfo", {}).get("loading"),
        "status": d.get("status") or d.get("taskInfo", {}).get("status"),
        "progress": d.get("progress") or d.get("taskInfo", {}).get("progress"),
        "url_count": url_count,
        "url": url[0] if isinstance(url, list) and url else (url if isinstance(url, str) else ""),
    }


def create_images():
    run([LIBTV, "project", "use", PROJECT_UUID], timeout=60)
    scan_groups_to_state()
    rows_by_task = {name: read_rows(cfg["source_xlsx"]) for name, cfg in TASKS.items()}
    jobs = []
    with ThreadPoolExecutor(max_workers=6) as pool:
        for task_name, rows in rows_by_task.items():
            for row in rows:
                jobs.append(pool.submit(create_image, task_name, TASKS[task_name], row))
        for fut in as_completed(jobs):
            print("[image]", fut.result(), flush=True)


def create_video(task_name, row):
    state = load_state()
    shot_key = str(row["shot"])
    if (task_name, shot_key) in SKIP_VIDEO:
        return task_name, row["shot"], "", "skipped"
    item = state["tasks"].setdefault(task_name, {}).setdefault(shot_key, {})
    if item.get("video_node"):
        return task_name, row["shot"], item["video_node"], "exists"
    image_node = item.get("image_node")
    if not image_node:
        raise RuntimeError(f"Missing image node for {task_name} shot {shot_key}")
    st = image_status(image_node)
    if st["url_count"] < 1:
        return task_name, row["shot"], "", "image-not-ready"
    duration = int(item.get("duration") or estimate_duration(row["line"], row["time"]))
    prompt = video_prompt(row, duration)
    cmd = [
        LIBTV,
        "node",
        "create",
        f"{task_name}_shot{row['shot']:02d}_video",
        "-t",
        "video",
        "--prompt",
        prompt,
        "-s",
        "model=Seedance 2.0 VIP",
        "-s",
        "modeType=singleImage2video",
        "-s",
        "ratio=9:16",
        "-s",
        "resolution=480p",
        "-s",
        f"duration={duration}",
        "-s",
        "enableSound=on",
        "--left",
        image_node,
    ]
    out = run_allow_created(cmd, timeout=300)
    data = first_json(out)
    node_id = data.get("newNodeKey") or data.get("nodeKey") or data.get("data", {}).get("nodeKey") or data.get("id")
    if not node_id:
        node_id = data.get("reactFlowNode", {}).get("id") or data.get("data", {}).get("key")
    if not node_id:
        raise RuntimeError(f"Could not parse video node id for {task_name} shot {shot_key}: {out[:1000]}")
    run_allow_created([LIBTV, "group", TASKS[task_name]["group_id"], "--node", node_id], timeout=120)
    detail = first_json(run([LIBTV, "node", "-p", PROJECT_UUID, node_id], timeout=90))
    prompt_readback = json.dumps(detail.get("data", {}).get("params", {}), ensure_ascii=False)
    if any(marker in prompt_readback for marker in JUNK_MARKERS):
        raise RuntimeError(f"Video prompt encoding failed for {task_name} shot {shot_key} node {node_id}")
    update_shot_state(task_name, shot_key, {"video_node": node_id, "duration": duration})
    return task_name, row["shot"], node_id, "created"


def scan_video_nodes_to_state():
    state = scan_groups_to_state()
    for task_name, cfg in TASKS.items():
        group = first_json(run([LIBTV, "group", cfg["group_id"]], timeout=90))
        ids = list(dict.fromkeys(group.get("executionPlan", []) + group.get("childNodeIds", [])))
        for node_id in ids:
            try:
                detail = first_json(run([LIBTV, "node", "-p", PROJECT_UUID, node_id], timeout=90))
            except Exception:
                continue
            data = detail.get("data", {})
            name = data.get("displayName") or data.get("name") or ""
            m = re.match(rf"{re.escape(task_name)}_shot(\d+)_video$", name)
            if not m:
                continue
            shot_key = str(int(m.group(1)))
            state["tasks"].setdefault(task_name, {}).setdefault(shot_key, {})["video_node"] = node_id
    save_state(state)
    return state


def create_videos():
    run([LIBTV, "project", "use", PROJECT_UUID], timeout=60)
    scan_video_nodes_to_state()
    rows_by_task = {name: read_rows(cfg["source_xlsx"]) for name, cfg in TASKS.items()}
    jobs = []
    with ThreadPoolExecutor(max_workers=6) as pool:
        for task_name, rows in rows_by_task.items():
            for row in rows:
                jobs.append(pool.submit(create_video, task_name, row))
        for fut in as_completed(jobs):
            print("[video]", fut.result(), flush=True)


def create_videos_fast():
    run([LIBTV, "project", "use", PROJECT_UUID], timeout=60)
    rows_by_task = {name: read_rows(cfg["source_xlsx"]) for name, cfg in TASKS.items()}
    jobs = []
    with ThreadPoolExecutor(max_workers=1) as pool:
        for task_name, rows in rows_by_task.items():
            for row in rows:
                jobs.append(pool.submit(create_video, task_name, row))
        for fut in as_completed(jobs):
            try:
                print("[video-fast]", fut.result(), flush=True)
            except Exception as exc:
                print("[video-fast-error]", str(exc)[:300], flush=True)


def trigger_image_node(task_name, shot, node_id):
    st = image_status(node_id)
    if st["url_count"] >= 1:
        return task_name, shot, node_id, "ready"
    if st.get("loading") or st.get("status") == 1:
        return task_name, shot, node_id, "running"
    try:
        run_allow_created([LIBTV, "node", "-p", PROJECT_UUID, node_id, "-r"], timeout=900)
    except subprocess.TimeoutExpired:
        return task_name, shot, node_id, "trigger-timeout-check-later"
    except Exception as exc:
        return task_name, shot, node_id, f"trigger-error:{str(exc)[:120]}"
    return task_name, shot, node_id, "triggered"


def video_status(node_id):
    data = first_json(run([LIBTV, "node", "-p", PROJECT_UUID, node_id], timeout=90))
    d = data.get("data", {})
    url = d.get("url")
    url_count = len(url) if isinstance(url, list) else (1 if url else 0)
    return {
        "loading": d.get("loading") or d.get("taskInfo", {}).get("loading"),
        "status": d.get("status") or d.get("taskInfo", {}).get("status"),
        "progress": d.get("progress") or d.get("taskInfo", {}).get("progress"),
        "url_count": url_count,
        "url": url[0] if isinstance(url, list) and url else (url if isinstance(url, str) else ""),
    }


def trigger_video_node(task_name, shot, node_id):
    st = video_status(node_id)
    if st["url_count"] >= 1:
        return task_name, shot, node_id, "ready"
    if st.get("loading") or st.get("status") == 1:
        return task_name, shot, node_id, "running"
    try:
        run_allow_created([LIBTV, "node", "-p", PROJECT_UUID, node_id, "-r"], timeout=1200)
    except subprocess.TimeoutExpired:
        return task_name, shot, node_id, "trigger-timeout-check-later"
    except Exception as exc:
        return task_name, shot, node_id, f"trigger-error:{str(exc)[:120]}"
    return task_name, shot, node_id, "triggered"


def trigger_videos():
    run([LIBTV, "project", "use", PROJECT_UUID], timeout=60)
    state = scan_video_nodes_to_state()
    jobs = []
    with ThreadPoolExecutor(max_workers=1) as pool:
        for task_name, shots in state.get("tasks", {}).items():
            for shot, item in sorted(shots.items(), key=lambda x: int(x[0])):
                if (task_name, shot) in SKIP_VIDEO:
                    continue
                node_id = item.get("video_node")
                if node_id:
                    jobs.append(pool.submit(trigger_video_node, task_name, int(shot), node_id))
        for fut in as_completed(jobs):
            print("[trigger-video]", fut.result(), flush=True)


def trigger_videos_fast():
    run([LIBTV, "project", "use", PROJECT_UUID], timeout=60)
    state = load_state()
    jobs = []
    with ThreadPoolExecutor(max_workers=6) as pool:
        for task_name, shots in state.get("tasks", {}).items():
            for shot, item in sorted(shots.items(), key=lambda x: int(x[0])):
                node_id = item.get("video_node")
                if node_id:
                    jobs.append(pool.submit(trigger_video_node, task_name, int(shot), node_id))
        for fut in as_completed(jobs):
            try:
                print("[trigger-video-fast]", fut.result(), flush=True)
            except Exception as exc:
                print("[trigger-video-fast-error]", str(exc)[:300], flush=True)


def poll_videos(max_minutes=180):
    deadline = time.time() + max_minutes * 60
    while True:
        state = load_state()
        all_items = []
        for task_name, shots in state.get("tasks", {}).items():
            for shot, item in shots.items():
                if (task_name, shot) in SKIP_VIDEO:
                    continue
                if item.get("video_node"):
                    all_items.append((task_name, shot, item["video_node"]))
        ready = 0
        for task_name, shot, node_id in all_items:
            st = video_status(node_id)
            state["tasks"][task_name][shot]["video_status"] = st
            if st["url_count"] >= 1:
                ready += 1
        save_state(state)
        print(f"[poll-videos] ready {ready}/{len(all_items)}", flush=True)
        if ready == len(all_items):
            return True
        if time.time() > deadline:
            return False
        time.sleep(60)


def trigger_images():
    run([LIBTV, "project", "use", PROJECT_UUID], timeout=60)
    state = scan_groups_to_state()
    jobs = []
    with ThreadPoolExecutor(max_workers=10) as pool:
        for task_name, shots in state.get("tasks", {}).items():
            for shot, item in sorted(shots.items(), key=lambda x: int(x[0])):
                node_id = item.get("image_node")
                if node_id:
                    jobs.append(pool.submit(trigger_image_node, task_name, int(shot), node_id))
        for fut in as_completed(jobs):
            print("[trigger-image]", fut.result(), flush=True)


def poll_images(max_minutes=90):
    deadline = time.time() + max_minutes * 60
    while True:
        state = load_state()
        all_items = []
        for task_name, shots in state.get("tasks", {}).items():
            for shot, item in shots.items():
                if item.get("image_node"):
                    all_items.append((task_name, shot, item["image_node"]))
        ready = 0
        for task_name, shot, node_id in all_items:
            st = image_status(node_id)
            state["tasks"][task_name][shot]["image_status"] = st
            if st["url_count"] >= 1 and st["status"] == 2:
                ready += 1
        save_state(state)
        print(f"[poll-images] ready {ready}/{len(all_items)}", flush=True)
        if ready == len(all_items):
            return True
        if time.time() > deadline:
            return False
        time.sleep(60)


def main():
    step = sys.argv[1] if len(sys.argv) > 1 else "images"
    if step == "images":
        create_images()
    elif step == "videos":
        create_videos()
    elif step == "videos-fast":
        create_videos_fast()
    elif step == "scan":
        state = scan_groups_to_state()
        print(json.dumps(state, ensure_ascii=False, indent=2))
    elif step == "trigger-images":
        trigger_images()
    elif step == "trigger-videos":
        trigger_videos()
    elif step == "trigger-videos-fast":
        trigger_videos_fast()
    elif step == "poll-images":
        ok = poll_images()
        raise SystemExit(0 if ok else 2)
    elif step == "poll-videos":
        ok = poll_videos()
        raise SystemExit(0 if ok else 2)
    elif step == "rows":
        for name, cfg in TASKS.items():
            rows = read_rows(cfg["source_xlsx"])
            print(name, len(rows), [r["shot"] for r in rows])
            for row in rows:
                print(row["shot"], estimate_duration(row["line"], row["time"]), row["line"])
    else:
        raise SystemExit(f"unknown step: {step}")


if __name__ == "__main__":
    main()
