import json
import subprocess
import sys
import time
import urllib.request
from pathlib import Path

import importlib.util
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RUNNER_PATH = Path(r"E:\灵鹤芝谷素材库\灵鹤芝谷工具矩阵\loops\lingzhi_tea_storyboard_video_loop\scripts\run_antioxidant_two_batch.py")
DELIVERY_ROOT = Path(r"E:\灵鹤芝谷素材库\灵鹤芝谷工具矩阵\outputs\批量脚本视频任务_0623\delivery_by_script")
LINK_STATE = DELIVERY_ROOT / "feishu_sheet_links.json"
LARK_CLI = r"C:\Users\gba\.workbuddy\binaries\node\versions\22.22.2\lark-cli.cmd"

spec = importlib.util.spec_from_file_location("runner", RUNNER_PATH)
runner = importlib.util.module_from_spec(spec)
spec.loader.exec_module(runner)


HEADERS = [
    "脚本",
    "镜头",
    "状态",
    "时长(秒)",
    "景别",
    "运镜",
    "分镜画面",
    "口播稿",
    "分镜图",
    "分镜图链接",
    "视频链接",
]

MOTION_MAP = {
    "slight handheld UGC camera movement": "轻微手持UGC运镜",
    "slow natural push-in": "缓慢推进",
    "smooth lateral scan": "平稳横移扫拍",
    "short-video quick-cut rhythm": "短视频快切节奏",
    "clean final still-frame composition": "干净定格收尾",
}


def run(cmd, cwd=None, timeout=900):
    proc = subprocess.run(
        cmd,
        cwd=str(cwd) if cwd else None,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        timeout=timeout,
        shell=False,
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stdout + "\n" + proc.stderr)
    return proc.stdout


def first_json(text):
    start = text.find("{")
    if start < 0:
        raise ValueError(text[:1000])
    depth = 0
    in_str = False
    esc = False
    for i, ch in enumerate(text[start:], start):
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return json.loads(text[start : i + 1])
    raise ValueError("unbalanced json")


def load_links():
    if LINK_STATE.exists():
        return json.loads(LINK_STATE.read_text(encoding="utf-8"))
    return {}


def save_links(state):
    LINK_STATE.parent.mkdir(parents=True, exist_ok=True)
    tmp = LINK_STATE.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(LINK_STATE)


def node_url(status):
    url = status.get("url") or ""
    return url if isinstance(url, str) else ""


def public_title(task_name):
    if task_name == "antioxidant1":
        return "抗氧化1_灵芝黄芪枸杞茶_分镜视频表"
    if task_name == "antioxidant4":
        return "抗氧化4_灵芝黄芪枸杞茶_分镜视频表"
    if task_name.startswith("script"):
        return f"脚本{int(task_name.replace('script', '')):02d}_灵芝黄芪枸杞茶_分镜视频表"
    return f"{task_name}_灵芝黄芪枸杞茶_分镜视频表"


def script_label(task_name):
    if task_name == "antioxidant1":
        return "抗氧化1"
    if task_name == "antioxidant4":
        return "抗氧化4"
    if task_name.startswith("script"):
        return f"脚本{int(task_name.replace('script', '')):02d}"
    return task_name


def refresh_items(task_name):
    cfg = runner.TASKS[task_name]
    rows = runner.read_rows(cfg["source_xlsx"])
    state = runner.load_state()
    shots = state.get("tasks", {}).get(task_name, {})
    items = []
    for row in rows:
        item = shots.get(str(row["shot"]), {})
        image_node = item.get("image_node")
        video_node = item.get("video_node")
        image_status = runner.image_status(image_node) if image_node else {"url_count": 0, "url": ""}
        external_video_path = item.get("external_video_path")
        external_video_url = item.get("external_video_url") or ""
        if external_video_path and Path(external_video_path).exists():
            video_status = {
                "url_count": 1,
                "url": external_video_url,
                "local_path": external_video_path,
                "provider": item.get("external_video_provider", "external"),
            }
        else:
            video_status = runner.video_status(video_node) if video_node else {"url_count": 0, "url": ""}
        duration = int(item.get("duration") or runner.estimate_duration(row["line"], row["time"]))
        items.append(
            {
                "row": row,
                "duration": duration,
                "image_url": node_url(image_status),
                "video_url": node_url(video_status),
                "video_local_path": video_status.get("local_path", ""),
                "image_ready": image_status.get("url_count", 0) >= 1,
                "video_ready": video_status.get("url_count", 0) >= 1,
            }
        )
    return items


def download(url, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.stat().st_size > 0:
        return
    tmp = path.with_suffix(path.suffix + ".tmp")
    urllib.request.urlretrieve(url, tmp)
    tmp.replace(path)


def build_xlsx(task_name, items):
    script_dir = DELIVERY_ROOT / task_name
    image_dir = script_dir / "images"
    video_dir = script_dir / "videos"
    table_path = script_dir / f"{task_name}_table.xlsx"
    script_dir.mkdir(parents=True, exist_ok=True)

    for item in items:
        shot = int(item["row"]["shot"])
        download(item["image_url"], image_dir / f"{task_name}_shot{shot:02d}.png")
        video_path = video_dir / f"{task_name}_shot{shot:02d}.mp4"
        local_video = item.get("video_local_path")
        if local_video and Path(local_video).exists():
            video_path.parent.mkdir(parents=True, exist_ok=True)
            if not video_path.exists() or video_path.stat().st_size == 0:
                video_path.write_bytes(Path(local_video).read_bytes())
        else:
            download(item["video_url"], video_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "storyboard"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, item in enumerate(items, start=2):
        row = item["row"]
        shot = int(row["shot"])
        image_path = image_dir / f"{task_name}_shot{shot:02d}.png"
        ws.cell(idx, 1, script_label(task_name))
        ws.cell(idx, 2, shot)
        ws.cell(idx, 3, "已完成")
        ws.cell(idx, 4, int(item["duration"]))
        ws.cell(idx, 5, row.get("shot_size", ""))
        motion = runner.motion_en((row.get("shot_size") or "") + (row.get("visual") or ""))
        ws.cell(idx, 6, MOTION_MAP.get(motion, "轻微手持UGC运镜"))
        ws.cell(idx, 7, row.get("visual", ""))
        ws.cell(idx, 8, row.get("line", ""))
        if image_path.exists():
            img = XLImage(str(image_path))
            img.width = 72
            img.height = 128
            ws.add_image(img, ws.cell(idx, 9).coordinate)
        img_cell = ws.cell(idx, 10, "查看分镜图")
        img_cell.hyperlink = item["image_url"]
        img_cell.style = "Hyperlink"
        vid_cell = ws.cell(idx, 11, "查看视频")
        vid_cell.hyperlink = item["video_url"]
        vid_cell.style = "Hyperlink"
        ws.row_dimensions[idx].height = 100
        for col in range(1, len(HEADERS) + 1):
            ws.cell(idx, col).alignment = Alignment(vertical="center", wrap_text=True)

    for idx, width in enumerate([12, 8, 10, 12, 16, 24, 50, 50, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(table_path)
    return table_path


def import_xlsx(task_name, table_path):
    out = run(
        [
            LARK_CLI,
            "drive",
            "+import",
            "--file",
            f".\\{table_path.name}",
            "--type",
            "sheet",
            "--name",
            public_title(task_name),
            "--as",
            "user",
            "--format",
            "json",
        ],
        cwd=table_path.parent,
        timeout=1200,
    )
    data = first_json(out)
    payload = data.get("data", data)
    token = payload.get("token") or payload.get("file_token")
    url = payload.get("url") or (f"https://ucnscwivbsrz.feishu.cn/sheets/{token}" if token else "")
    if not token or not url:
        raise RuntimeError(f"import result missing token/url: {out[:1000]}")
    return {"token": token, "url": url, "title": public_title(task_name)}


def set_permission(token):
    data = {
        "link_share_entity": "tenant_readable",
        "share_entity": "same_tenant",
        "security_entity": "anyone_can_view",
    }
    run(
        [
            LARK_CLI,
            "drive",
            "permission.public",
            "patch",
            "--token",
            token,
            "--type",
            "sheet",
            "--data",
            json.dumps(data, ensure_ascii=False),
            "--as",
            "user",
            "--yes",
            "--format",
            "json",
        ],
        timeout=300,
    )


def deliver(task_name):
    items = refresh_items(task_name)
    total = len(items)
    img = sum(1 for x in items if x["image_ready"])
    vid = sum(1 for x in items if x["video_ready"])
    if img != total or vid != total:
        print(f"[pending] {task_name} images={img}/{total} videos={vid}/{total}", flush=True)
        return None
    table_path = build_xlsx(task_name, items)
    links = load_links()
    if task_name in links and links[task_name].get("url"):
        print(f"[exists] {task_name} {links[task_name]['url']}", flush=True)
        return links[task_name]
    result = import_xlsx(task_name, table_path)
    links[task_name] = result
    save_links(links)
    try:
        set_permission(result["token"])
        links[task_name]["permission"] = "tenant_readable"
        save_links(links)
    except Exception as exc:
        links[task_name]["permission_error"] = str(exc)[:500]
        save_links(links)
        print(f"[permission-error] {task_name} {exc}", flush=True)
    print(f"[delivered] {task_name} {result['url']}", flush=True)
    time.sleep(1)
    return result


tasks = sys.argv[1:] or ["antioxidant1", "antioxidant4"]
for task in tasks:
    deliver(task)
